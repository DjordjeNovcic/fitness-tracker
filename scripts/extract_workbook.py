from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = Path("/Users/dnovcic/Downloads/Jelovnik i kalorije - unapredjeno.xlsx")
OUTPUT_PATH = ROOT / "data" / "seed-data.js"

DAYS = {
    "PONEDELJAK": "Ponedeljak",
    "UTORAK": "Utorak",
    "SREDA": "Sreda",
    "CETVRTAK": "Cetvrtak",
    "PETAK": "Petak",
    "SUBOTA": "Subota",
    "NEDELJA": "Nedelja",
}


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def safe_number(value) -> float:
    if value in (None, ""):
        return 0.0
    return round(float(value), 4)


def parse_foods(workbook) -> list[dict]:
    sheet = workbook["Namirnice"]
    foods: list[dict] = []

    for row in range(3, sheet.max_row + 1):
        name = sheet.cell(row, 2).value
        if not name:
            continue
        food_id = int(sheet.cell(row, 1).value)
        foods.append(
            {
                "id": f"food-{food_id}",
                "name": str(name).strip(),
                "servingBaseGrams": safe_number(sheet.cell(row, 3).value) or 100,
                "kcal": safe_number(sheet.cell(row, 4).value),
                "protein": safe_number(sheet.cell(row, 5).value),
                "carbs": safe_number(sheet.cell(row, 6).value),
                "fat": safe_number(sheet.cell(row, 7).value),
                "category": str(sheet.cell(row, 8).value or "Ostalo").strip(),
            }
        )

    return foods


def parse_profile(workbook) -> dict:
    sheet = workbook["Jelovnik"]
    raw_weight = str(sheet["H2"].value or "").replace("kg", "").strip()
    return {
        "name": str(sheet["B2"].value or "").strip() or "Moj profil",
        "age": int(float(sheet["E2"].value or 0)) if sheet["E2"].value else 0,
        "weightKg": safe_number(raw_weight),
    }


def parse_goals(workbook) -> dict:
    sheet = workbook["Makronutrijenti"]
    return {
        "calories": round(safe_number(sheet["D5"].value)),
        "protein": safe_number(sheet["B2"].value),
        "carbs": safe_number(sheet["B3"].value),
        "fat": safe_number(sheet["B4"].value),
    }


def parse_weekly_plan(workbook, foods: list[dict]) -> list[dict]:
    sheet = workbook["Jelovnik"]
    food_id_by_name = {food["name"]: food["id"] for food in foods}
    entries: list[dict] = []
    current_day = None
    current_meal = None
    counter = 1

    for row in range(1, sheet.max_row + 1):
        label = sheet.cell(row, 1).value
        grams = sheet.cell(row, 3).value

        if isinstance(label, str):
            normalized = label.strip().upper()
            if normalized in DAYS:
                current_day = DAYS[normalized]
                current_meal = None
                continue
            if re.match(r"^\d+\.\s", label.strip()):
                current_meal = label.strip()
                continue
            if label.startswith("Namirnica") or label.startswith("Ukupno") or label.startswith("Personalne"):
                continue

        if not current_day or not current_meal or not label or not isinstance(grams, (int, float)) or grams <= 0:
            continue

        entries.append(
            {
                "id": f"plan-{counter}",
                "weekday": current_day,
                "mealLabel": current_meal,
                "foodId": food_id_by_name.get(str(label).strip()),
                "foodName": str(label).strip(),
                "grams": safe_number(grams),
            }
        )
        counter += 1

    return entries


def parse_training(workbook) -> list[dict]:
    sheet = workbook["Trening"]
    exercises: list[dict] = []

    for row in range(2, sheet.max_row + 1):
        raw = sheet.cell(row, 1).value
        if not raw:
            continue
        label = str(raw).strip()
        name = re.split(r"\s+\d", label, maxsplit=1)[0].strip()
        exercises.append(
            {
                "id": f"exercise-{slugify(label)}",
                "name": name or label,
                "details": label,
            }
        )

    return [
        {
            "id": "template-osnovni",
            "weekday": "Ponedeljak",
            "name": "Osnovni trening",
            "exercises": exercises,
        }
    ]


def parse_measurements(workbook) -> dict:
    sheet = workbook["Deficit pracenje"]
    headers = []
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(1, column).value
        if value:
            headers.append(str(value).strip().rstrip(":"))
    return {"trackedFields": headers}


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    foods = parse_foods(workbook)
    payload = {
        "version": 1,
        "profile": parse_profile(workbook),
        "goals": parse_goals(workbook),
        "foods": foods,
        "weeklyPlanEntries": parse_weekly_plan(workbook, foods),
        "trainingTemplates": parse_training(workbook),
        "meta": parse_measurements(workbook),
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        "window.SEED_DATA = " + json.dumps(payload, ensure_ascii=True, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
